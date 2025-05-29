from io import BytesIO
from datetime import datetime, date, time
from logging import getLogger
from functools import partial
from typing import Annotated, Callable, Union, TypeAlias
from fastapi import Depends, HTTPException, status, UploadFile, BackgroundTasks
from fastapi.routing import APIRouter
from enum import StrEnum
import openpyxl
from zipfile import BadZipFile

from app import auth
from app.admin.models import VendorId
from app.db import DB_V2, Session
from app.v2.models import *
from app.v2.pricing import transform, fetch_pricing
from app.v2.routes.vendor_product_class_discounts import (
    new_vendor_product_class_discount,
    mod_vendor_product_class_discount,
)
from app.v2.routes.vendor_pricing_by_customer import (
    mod_vendor_pricing_by_customer,
    new_vendor_pricing_by_customer,
)
from app.v2.routes.vendor_product_discounts import (
    mod_vendor_product_discount,
    new_vendor_product_discount,
)

from app.admin.models import (
    VendorId,
    FullPricingWithLink,
    PriceTemplateSheetColumns,
    PriceTemplateSheet,
    PriceTemplateModels,
    ProductCategoryDiscount as ProductCategoryDiscountDTO,
    ProductDiscount as ProductDiscountDTO,
    CustomerPrice as CustomerPriceDTO,
)
from app.downloads import (
    XLSXFileResponse,
    FileResponse,
    DownloadIDs,
    StreamingResponse,
)
from app.jsonapi.sqla_models import Vendor
from app.adp.utils.workbook_factory import generate_program

PARENT_PREFIX = "/v2/vendors"
VENDOR_PREFIX = "/{vendor}"
VENDORS = Vendor.__jsonapi_type_override__

logger = getLogger("uvicorn.info")
pricing = APIRouter(prefix="", tags=["v2"])


def date_to_datetime(d: date) -> Optional[datetime]:
    if not d:
        return
    return datetime.combine(d, time(0, 0))


Token = Annotated[auth.VerifiedToken, Depends(auth.authenticate_auth0_token)]
NewSession = Annotated[Session, Depends(DB_V2.get_db)]


class ReturnType(StrEnum):
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"


def generate_pricing_dl_link(
    vendor_id: VendorId, customer_id: int, callback: Callable
) -> str:
    resource_path = (
        f"/v2/vendors/{vendor_id}/vendor-customers/{customer_id}/pricing/download"
    )
    download_id = DownloadIDs.add_request(resource=resource_path, callback=callback)
    query = f"?download_id={download_id}"
    link = resource_path + query
    return link


@pricing.get(
    "/{vendor_id}/vendor-customers/{customer_id}/pricing",
    response_model=FullPricingWithLink,
    response_model_exclude_none=True,
    tags=["vendor customers", "pricing"],
)
async def vendor_customer_pricing(
    token: Token,
    session: NewSession,
    vendor_id: VendorId,
    customer_id: int,
    effective_date: date = None,
    return_type: ReturnType = ReturnType.JSON,
) -> FullPricingWithLink:
    """
    Getting pricing can be challenging to generalize on the front end, so logic here
    will do special method routing by-vendor if the request passes the auth
    check.

    Non-default return_types set in the query will still return JSON but only
    the download_link (no JSON pricing object).

    Execution of pricing file generation in this case is deferred
    to the return process in the route where the link is redeemed.
    """
    try:
        customer = (
            auth.VendorCustomerOperations(token, VendorCustomer, id=vendor_id)
            .allow_admin()
            .allow_sca()
            .allow_dev()
            .allow_customer("std")
            .get(session, obj_id=customer_id)
        )
    except HTTPException as e:
        raise e

    if effective_date:
        effective_date = date_to_datetime(effective_date)
    price_fetch = partial(fetch_pricing, session, vendor_id, customer_id)
    transform_ = partial(transform, customer, vendor_id, effective_date=effective_date)
    match vendor_id, return_type:
        # ReturnType.JSON: return pricing along with a download link to a CSV file
        # ReturnType.CSV: return a download link with deferred execution
        # ReturnType.XLSX: return a download link with deferred execution
        case VendorId.ATCO, ReturnType.JSON:
            remove_cols = ["fp_ean", "upc_code"]
            pricing = price_fetch(mode="both")
            cb = partial(transform_, pricing, remove_cols)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link, pricing=pricing)

        case VendorId.ATCO, (ReturnType.CSV | ReturnType.XLSX):
            remove_cols = ["fp_ean", "upc_code"]
            pricing = partial(price_fetch, mode="both")
            cb = partial(transform_, pricing, remove_cols, file_type=return_type.value)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link)

        case VendorId.ADP, ReturnType.JSON:
            remove_cols = None
            pricing = price_fetch(mode="customer")
            # ADP uses a special file generation method
            cb = partial(
                generate_program,
                session=session,
                customer_id=customer_id,
                effective_date=effective_date,
            )
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link, pricing=pricing)

        case VendorId.ADP, ReturnType.XLSX:
            remove_cols = None
            # pricing = partial(fetch_pricing, mode="customer")
            # ADP uses a special file generation method
            cb = partial(
                generate_program,
                session=session,
                customer_id=customer_id,
                effective_date=effective_date,
            )
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link)

        case VendorId.VYBOND, ReturnType.JSON:
            keys_to_override = ["KEY", "STATE_CPD"]
            remove_cols = ["ucc", "upc"]
            pricing = price_fetch(mode="both", categories_to_override=keys_to_override)
            pivot = False
            cb = partial(transform_, pricing, remove_cols, pivot)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link, pricing=pricing)

        case VendorId.VYBOND, (ReturnType.CSV | ReturnType.XLSX):
            keys_to_override = ["KEY", "STATE_CPD"]
            remove_cols = ["ucc", "upc"]
            pricing = partial(
                price_fetch, mode="both", categories_to_override=keys_to_override
            )
            pivot = False
            cb = partial(transform_, pricing, remove_cols, pivot, file_type=return_type)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link)

        case VendorId.FRIEDRICH, ReturnType.JSON:
            remove_cols = None
            pricing = price_fetch(mode="both")
            pivot = False
            cb = partial(transform_, pricing, remove_cols, pivot)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link, pricing=pricing)

        case VendorId.FRIEDRICH, (ReturnType.CSV | ReturnType.XLSX):
            remove_cols = None
            pricing = partial(price_fetch, mode="both")
            pivot = False
            cb = partial(transform_, pricing, remove_cols, pivot, file_type=return_type)
            dl_link = generate_pricing_dl_link(vendor_id, customer_id, cb)
            return FullPricingWithLink(download_link=dl_link)

        case _:
            raise HTTPException(status.HTTP_501_NOT_IMPLEMENTED)


def create_api_body_from_DTO(
    dto_model: BaseModel,
    vendor_id: VendorId,
    customer_id: int,
    effective_date: datetime,
    session: Session,
) -> tuple[Union[BaseModel, dict], bool]:
    """dispatch"""
    if isinstance(dto_model, CustomerPriceDTO):
        return create_vendor_pricing_by_customer_body(
            dto_model, session, vendor_id, customer_id, effective_date
        )
    elif isinstance(dto_model, ProductCategoryDiscountDTO):
        return create_vendor_product_class_discount_body(
            dto_model, session, vendor_id, customer_id, effective_date
        )
    elif isinstance(dto_model, ProductDiscountDTO):
        return create_vendor_product_discount_body(
            dto_model, session, vendor_id, customer_id, effective_date
        )


def create_vendor_pricing_by_customer_body(
    record_cp: CustomerPriceDTO,
    session: Session,
    vendor_id: VendorId,
    customer_id: int,
    effective_date: datetime,
) -> tuple[Union[NewVendorPricingByCustomerRObj, dict], bool]:
    try:
        price_category = record_cp.get_price_category_id(session, vendor_id)
        product = record_cp.get_product_id(session, vendor_id)
        data = NewVendorPricingByCustomerRObj(
            type="vendor-pricing-by-customer",
            attributes=VendorPricingByCustomerAttrs(
                price=int(record_cp.price * 100),
                effective_date=effective_date,
                use_as_override=record_cp.is_override,
            ),
            relationships=VendorPricingByCustomerRels(
                vendors=JSONAPIRelationships(
                    data=[JSONAPIResourceIdentifier(id=vendor_id.value, type="vendors")]
                ),
                vendor_customers=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=customer_id, type="vendor-customers"
                        )
                    ]
                ),
                vendor_products=JSONAPIRelationships(
                    data=[JSONAPIResourceIdentifier(id=product, type="vendor-products")]
                ),
                vendor_pricing_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=price_category, type="vendor-pricing-classes"
                        )
                    ]
                ),
            ),
        )
    except Exception as e:
        return {"record": record_cp.model_dump(), "error": e}, True
    else:
        return data, False


def create_vendor_product_class_discount_body(
    record_pcd: ProductCategoryDiscountDTO,
    session: Session,
    vendor_id: VendorId,
    customer_id: int,
    effective_date: datetime,
) -> tuple[Union[NewVendorProductClassDiscountRObj, dict], bool]:
    try:
        # create
        base_class_id = record_pcd.get_base_price_category_id(session, vendor_id)
        label_class_id = record_pcd.get_label_price_category_id(session, vendor_id)
        product_class_id = record_pcd.get_product_category_id(session, vendor_id)
        data = NewVendorProductClassDiscountRObj(
            type="vendor-product-class-discounts",
            attributes=VendorProductClassDiscountAttrs(
                discount=record_pcd.discount,
                effective_date=effective_date,
            ),
            relationships=VendorProductClassDiscountRels(
                vendors=JSONAPIRelationships(
                    data=[JSONAPIResourceIdentifier(id=vendor_id.value, type="vendors")]
                ),
                vendor_customers=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=customer_id, type="vendor-customers"
                        )
                    ]
                ),
                base_price_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=base_class_id,
                            type="base-price-classes",
                        )
                    ]
                ),
                label_price_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=label_class_id,
                            type="label-price-classes",
                        )
                    ]
                ),
                vendor_product_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=product_class_id,
                            type="vendor-product-classes",
                        )
                    ]
                ),
            ),
        )
    except Exception as e:
        return {"record": record_pcd.model_dump(), "error": e}, True
    else:
        return data, False


def create_vendor_product_discount_body(
    record_pd: ProductDiscountDTO,
    session: Session,
    vendor_id: VendorId,
    customer_id: int,
    effective_date: datetime,
) -> tuple[Union[NewVendorProductDiscountRObj, dict], bool]:
    try:
        # create
        base_class_id = record_pd.get_base_price_category_id(session, vendor_id)
        label_class_id = record_pd.get_label_price_category_id(session, vendor_id)
        product_id = record_pd.get_product_id(session, vendor_id)
        data = NewVendorProductDiscountRObj(
            type="vendor-product-discounts",
            attributes=VendorProductDiscountAttrs(
                discount=record_pd.discount,
                effective_date=effective_date,
            ),
            relationships=VendorProductDiscountRels(
                vendors=JSONAPIRelationships(
                    data=[JSONAPIResourceIdentifier(id=vendor_id.value, type="vendors")]
                ),
                vendor_customers=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=customer_id, type="vendor-customers"
                        )
                    ]
                ),
                base_price_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=base_class_id,
                            type="base-price-classes",
                        )
                    ]
                ),
                label_price_classes=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=label_class_id,
                            type="label-price-classes",
                        )
                    ]
                ),
                vendor_products=JSONAPIRelationships(
                    data=[
                        JSONAPIResourceIdentifier(
                            id=product_id,
                            type="vendor-products",
                        )
                    ]
                ),
            ),
        )
    except Exception as e:
        return {"record": record_pd.model_dump(), "error": e}, True
    else:
        return data, False


async def perform_insert(
    token: Token,
    session: Session,
    route_method: Callable,
    data: BaseModel,
    response_class: type[BaseModel],
) -> tuple[Optional[BaseModel], bool]:
    try:
        new_record = await route_method(
            token,
            session,
            data,
        )
        new_record = response_class(**new_record)
    except HTTPException as http_e:
        if http_e.status_code == 409:
            return None, True
        else:
            raise http_e
    except Exception as e:
        raise e
    else:
        return new_record, False


async def fallback_update(
    token: Token,
    session: Session,
    route_method: Callable,
    data: BaseModel,
    response_class: type[BaseModel],
    id_lookup_sql: str,
    params: dict,
) -> tuple[Optional[BaseModel], bool]:
    existing_id = DB_V2.execute(session, id_lookup_sql, params).scalar_one()
    data.data.id = existing_id
    try:
        updated_record = await route_method(token, session, existing_id, data)
        updated_record = response_class(**updated_record)
    except Exception as e:
        logger.error(f"Error updating product discount: {e}")
        return {"record": data.model_dump(), "error": e}, True
    else:
        return updated_record, False


@pricing.post(
    "/{vendor_id}/vendor-customers/{customer_id}/pricing",
    response_model=FullPricingWithLink,
    response_model_exclude_none=True,
    tags=["vendor customers", "pricing"],
)
async def upsert_vendor_customer_pricing_from_file(
    token: Token,
    session: NewSession,
    vendor_id: VendorId,
    customer_id: int,
    templated_file: UploadFile,
    effective_date: date = datetime.today().date(),
) -> FullPricingWithLink:
    """Upsert customer pricing for a vendor."""
    try:
        customer = (
            auth.VendorCustomerOperations(token, VendorCustomer, id=vendor_id)
            .allow_admin()
            .allow_sca()
            .allow_dev()
            .get(session, obj_id=customer_id)
        )
    except HTTPException as e:
        raise e

    file_data = BytesIO(await templated_file.read())
    try:
        parsed_data = parse_pricing_template(file_data)
    except BadZipFile:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "The uploaded file is not loading correctly. Expecting an Excel file.",
        )
    MetaObj: TypeAlias = dict[str, list[dict[str, str]]]
    errors: MetaObj = {}
    updates: MetaObj = {}
    inserts: MetaObj = {}

    run_order = [
        PriceTemplateSheet.PRODUCT_CATEGORY_DISCOUNTS,
        PriceTemplateSheet.PRODUCT_DISCOUNTS,
        PriceTemplateSheet.CUSTOMER_PRICING,
    ]
    parsed_data_ordered = {k: parsed_data[k] for k in run_order if k in parsed_data}
    for record_type, records in parsed_data_ordered.items():
        # TODO the logic under each case is very similar
        # but I'm not sure if trying to abstract it is a good idea or not
        match record_type:
            case PriceTemplateSheet.CUSTOMER_PRICING:
                """set vendor_pricing_by_customer"""
                errors.setdefault(record_type, [])
                updates.setdefault(record_type, [])
                inserts.setdefault(record_type, [])
                for record_cp in records:
                    # create
                    record_cp: CustomerPriceDTO
                    customer_pricing_data: NewVendorPricingByCustomerRObj
                    customer_pricing_data, error = create_api_body_from_DTO(
                        record_cp,
                        vendor_id=vendor_id,
                        customer_id=customer_id,
                        effective_date=effective_date,
                        session=session,
                    )
                    if error:
                        errors[record_type].append(customer_pricing_data)
                        continue
                    try:
                        # insert
                        new_price: VendorPricingByCustomerResourceResp
                        new_price, conflict = await perform_insert(
                            token,
                            session,
                            new_vendor_pricing_by_customer,
                            NewVendorPricingByCustomer(data=customer_pricing_data),
                            VendorPricingByCustomerResourceResp,
                        )
                        if conflict:
                            # update
                            sql = """
                                SELECT id
                                FROM vendor_pricing_by_customer
                                WHERE vendor_customer_id = :vc_id
                                AND product_id = :p_id
                                AND pricing_class_id = :pc_id;
                            """
                            rels = customer_pricing_data.relationships
                            params = dict(
                                vc_id=customer_id,
                                p_id=rels.vendor_products.data[0].id,
                                pc_id=rels.vendor_pricing_classes.data[0].id,
                            )
                            mod_obj = ModVendorPricingByCustomerRObj(
                                id=0,
                                type=customer_pricing_data.type,
                                attributes=customer_pricing_data.attributes,
                                relationships=customer_pricing_data.relationships,
                            )
                            updated_record, error = await fallback_update(
                                token,
                                session,
                                mod_vendor_pricing_by_customer,
                                ModVendorPricingByCustomer(data=mod_obj),
                                VendorPricingByCustomerResourceResp,
                                sql,
                                params,
                            )
                            if error:
                                logger.error(f"Error updating customer price: {e}")
                                errors[record_type].append(updated_record)
                                continue
                            else:
                                updated_record: VendorPricingByCustomerResourceResp
                                updates[record_type].append(
                                    {
                                        "record": record_cp.model_dump(),
                                        "id": updated_record.data.id,
                                    }
                                )
                        else:
                            inserts[record_type].append(
                                {
                                    "record": record_cp.model_dump(),
                                    "id": new_price.data.id,
                                }
                            )
                    except Exception as e:
                        logger.error(f"Error establishing customer price: {e}")
                        errors[record_type].append(
                            {"record": record_cp.model_dump(), "error": e}
                        )
                        continue

            case PriceTemplateSheet.PRODUCT_CATEGORY_DISCOUNTS:
                """set product category discounts, which will set pricing as well
                as a side effect"""
                errors.setdefault(record_type, [])
                updates.setdefault(record_type, [])
                inserts.setdefault(record_type, [])
                for record_pcd in records:
                    record_pcd: ProductCategoryDiscountDTO
                    product_class_discount_data: NewVendorProductClassDiscountRObj
                    product_class_discount_data, error = create_api_body_from_DTO(
                        record_pcd,
                        vendor_id=vendor_id,
                        customer_id=customer_id,
                        effective_date=effective_date,
                        session=session,
                    )
                    if error:
                        errors[record_type].append(product_class_discount_data)
                        continue
                    try:
                        # insert
                        new_discount: VendorProductClassDiscountResourceResp
                        new_discount, conflict = await perform_insert(
                            token,
                            session,
                            new_vendor_product_class_discount,
                            NewVendorProductClassDiscount(
                                data=product_class_discount_data
                            ),
                            VendorProductClassDiscountResourceResp,
                        )
                        if conflict:
                            # update
                            ex_id_sql = """
                                SELECT id
                                FROM vendor_product_class_discounts
                                WHERE base_price_class = :bpc
                                AND label_price_class = :lpc
                                AND product_class_id = :pcid
                                AND vendor_customer_id = :vcid
                            """
                            rels = product_class_discount_data.relationships
                            params = dict(
                                bpc=rels.base_price_classes.data[0].id,
                                lpc=rels.label_price_classes.data[0].id,
                                pcid=rels.vendor_product_classes.data[0].id,
                                vcid=customer_id,
                            )
                            mod_obj = ModVendorProductClassDiscountRObj(
                                id=0,
                                type=product_class_discount_data.type,
                                attributes=product_class_discount_data.attributes,
                                relationships=product_class_discount_data.relationships,
                            )
                            updated_record, error = await fallback_update(
                                token,
                                session,
                                mod_vendor_product_class_discount,
                                ModVendorProductClassDiscount(data=mod_obj),
                                VendorProductClassDiscountResourceResp,
                                ex_id_sql,
                                params,
                            )

                            if error:
                                logger.error(f"Error updating product discount: {e}")
                                errors[record_type].append(updated_record)
                                continue
                            else:
                                updated_record: VendorProductClassDiscountResourceResp
                                updates[record_type].append(
                                    {
                                        "record": record_pcd.model_dump(),
                                        "id": updated_record.data.id,
                                    }
                                )
                        else:
                            inserts[record_type].append(
                                {
                                    "record": record_pcd.model_dump(),
                                    "id": new_discount.data.id,
                                }
                            )
                    except Exception as e:
                        logger.error(f"Error establishing product discount: {e}")
                        errors[record_type].append(
                            {"record": record_pcd.model_dump(), "error": e}
                        )
                        continue

            case PriceTemplateSheet.PRODUCT_DISCOUNTS:
                """set product-specific discounts, which will set pricing
                as a side effect"""
                errors.setdefault(record_type, [])
                updates.setdefault(record_type, [])
                inserts.setdefault(record_type, [])
                for record_pd in records:
                    record_pd: ProductDiscountDTO
                    product_discount_data: NewVendorProductDiscountRObj
                    product_discount_data, error = create_api_body_from_DTO(
                        record_pd,
                        vendor_id=vendor_id,
                        customer_id=customer_id,
                        effective_date=effective_date,
                        session=session,
                    )
                    if error:
                        errors[record_type].append(product_discount_data)
                        continue
                    try:
                        # insert
                        new_product_discount: VendorProductDiscountResourceResp
                        new_product_discount, conflict = await perform_insert(
                            token,
                            session,
                            new_vendor_product_discount,
                            NewVendorProductDiscount(data=product_discount_data),
                            VendorProductDiscountResourceResp,
                        )
                        if conflict:
                            # update
                            sql = """
                                SELECT id
                                FROM vendor_product_discounts
                                WHERE vendor_customer_id = :vc_id
                                AND product_id = :p_id
                                AND base_price_class = :bpc_id
                                AND label_price_class = :lpc_id;
                            """
                            rels = product_discount_data.relationships
                            params = dict(
                                vc_id=customer_id,
                                p_id=rels.vendor_products.data[0].id,
                                bpc_id=rels.base_price_classes.data[0].id,
                                lpc_id=rels.label_price_classes.data[0].id,
                            )
                            mod_obj = ModVendorProductDiscountRObj(
                                id=0,
                                type=product_discount_data.type,
                                attributes=product_discount_data.attributes,
                                relationships=product_discount_data.relationships,
                            )
                            updated_record, error = await fallback_update(
                                token,
                                session,
                                mod_vendor_product_discount,
                                ModVendorProductDiscount(data=mod_obj),
                                VendorProductDiscountResourceResp,
                                sql,
                                params,
                            )
                            if error:
                                logger.error(f"Error updating product discount: {e}")
                                errors[record_type].append(updated_record)
                                continue
                            else:
                                updated_record: VendorProductDiscountResourceResp
                                updates[record_type].append(
                                    {
                                        "record": record_pd.model_dump(),
                                        "id": updated_record.data.id,
                                    }
                                )
                        else:
                            inserts[record_type].append(
                                {
                                    "record": record_pd.model_dump(),
                                    "id": new_product_discount.data.id,
                                }
                            )
                    except Exception as e:
                        logger.error(f"Error establishing product discount: {e}")
                        errors[record_type].append(
                            {"record": record_pd.model_dump(), "error": e}
                        )
                        continue

    session.commit()
    new_meta = dict(errors=errors, updates=updates, inserts=inserts)
    ret = await vendor_customer_pricing(
        token,
        session,
        vendor_id,
        customer_id,
        None,
        ReturnType.JSON,
    )
    ret.meta = new_meta
    return ret


@pricing.get(
    "/{vendor_id}/vendor-customers/{customer_id}/pricing/download",
    response_class=StreamingResponse,
    response_model=None,
    tags=["vendor customers", "pricing", "download"],
)
async def download_price_file(
    vendor_id: VendorId, customer_id: int, download_id: str
) -> Union[FileResponse, XLSXFileResponse]:
    resource_path = (
        f"/v2/vendors/{vendor_id}/vendor-customers/{customer_id}/pricing/download"
    )
    dl_obj = DownloadIDs.use_download(
        resource=resource_path,
        id_value=download_id,
    )
    try:
        file = dl_obj.callback()
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        else:
            raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))

    else:
        return file


def parse_pricing_template(data: BytesIO) -> dict[StrEnum, list[BaseModel]]:
    """for each sheet in the data (Excel File), return collections of pydantic
    objects representing the data in each row for each sheet"""
    data.seek(0)
    wb = openpyxl.load_workbook(data, data_only=True)
    ret: dict[StrEnum, list] = {}
    try:
        sheet_names = set(wb.sheetnames)
        expected_sheets = set([sheet.value for sheet in PriceTemplateSheet])
        if extra_sheets := sheet_names - expected_sheets:
            logger.warning(f"Additional sheets present: {', '.join(extra_sheets)}")
        elif expected_sheets - sheet_names == expected_sheets:
            msg = "None of the expected sheets are present in the file."
            msg += " Expected at least one sheet with the following names: "
            msg += f"{', '.join(expected_sheets)}"
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=msg)
        elif missing_sheets := expected_sheets - sheet_names:
            logger.warning(f"Missing Sheets: {', '.join(missing_sheets)}")

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            template_sheet_enum = PriceTemplateSheet(sheet_name)
            expected_columns = set(PriceTemplateSheetColumns[template_sheet_enum])
            visited = set()
            for cell in sheet[1]:
                if cell.value not in expected_columns:
                    msg = f"Unexpected column in {sheet_name}: {cell.value}"
                    raise HTTPException(status.HTTP_400_BAD_REQUEST, msg)
                else:
                    visited.add(cell.value)
            if missing := expected_columns - visited:
                msg = f"Expected columns for sheet {sheet_name} are missing: {missing}"
                raise HTTPException(status.HTTP_400_BAD_REQUEST, msg)
            first_row = [cell.value for cell in sheet[2]]
            missing_or_incomplete_first_row = not any(first_row) or not all(first_row)
            if missing_or_incomplete_first_row:
                logger.warning(
                    f"Skipping {sheet_name}: none/incomplete data in first row"
                )
                continue
            else:
                data_model = PriceTemplateModels[template_sheet_enum]
                for row in sheet.iter_rows(min_row=2):
                    row_vals = (cell.value for cell in row)
                    incomplete_row = not any(row_vals) or not all(row_vals)
                    if incomplete_row:
                        continue
                    data_obj = {
                        k: v
                        for k, v in zip(
                            data_model.model_fields.keys(), (cell.value for cell in row)
                        )
                    }
                    ret.setdefault(template_sheet_enum, list())
                    ret[template_sheet_enum].append(data_model(**data_obj))
    except Exception as e:
        raise e
    else:
        ret = {k: v for k, v in ret.items() if v}
        if ret:
            return ret
        msg = "Pricing Update file is empty"
        raise HTTPException(status.HTTP_400_BAD_REQUEST, msg)
    finally:
        wb.close()
        data.seek(0)
