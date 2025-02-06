from dotenv import load_dotenv

load_dotenv()
import re
import os
from json import loads
from pathlib import Path
from copy import copy
from io import BytesIO
import pandas as pd
import openpyxl as opxl
from typing import Literal
from logging import getLogger
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from app.adp.utils.programs import CustomerProgram
from app.adp.adp_models.model_series import Fields
from app.adp.adp_models import MODELS, HE, HH, MH, V, F, S, B, CE, CF, ModelSeries
from app.db import Stage, S3

from sqlalchemy.orm import Session

NOMENCLATURE_COL_WIDTH = 20
NOMENCLATURES = os.getenv("NOMENCLATURES")
STATIC_DIR = Path(os.getenv("STATIC_DIR"))
logger = getLogger("uvicorn.info")

CellRange = tuple[tuple[Cell]]
LenOption = Literal["long", "med", "short"]
MinCoords = tuple[int | None, int | None]
DataByCategory = dict[str, pd.DataFrame]


class FileGenExecption(Exception): ...


class AnchorPosition:
    def __init__(self, anchor_cell: str, offset_x: int, offset_y: int) -> None:
        self._anchor_cell = anchor_cell
        self._offset_x = offset_x
        self._offset_y = offset_y
        coords = coordinate_from_string(anchor_cell)
        ## anchoring works off of 0-indexing, coordinates are 1-indexed
        col, row = column_index_from_string(coords[0]) - 1, coords[1] - 1
        x_emu = pixels_to_EMU(offset_x)
        y_emu = pixels_to_EMU(offset_y)
        self.anchor = AnchorMarker(col=col, colOff=x_emu, row=row, rowOff=y_emu)

    def __repr__(self) -> str:
        return (
            f"AnchorPosition(anchor_cell={self._anchor_cell}, "
            f"offset_x={self._offset_x}, offset_y={self._offset_y})"
        )


class Logo:
    def __init__(
        self,
        price_pos: AnchorPosition,
        ratings_pos: AnchorPosition,
        parts_pos: AnchorPosition,
        nomen_long_pos: AnchorPosition,
        nomen_med_pos: AnchorPosition,
        nomen_short_pos: AnchorPosition,
    ) -> None:
        self.name: str
        self.img: str | BytesIO
        self.price_pos = price_pos
        self.ratings_pos = ratings_pos
        self.parts_pos = parts_pos
        self.nomen_long_pos = nomen_long_pos
        self.nomen_med_pos = nomen_med_pos
        self.nomen_short_pos = nomen_short_pos

    def __str__(self) -> str:
        return self.name

    def __repr__(self) -> str:
        position_attrs = {k: v for k, v in self.__dict__.items() if k.endswith("pos")}
        arg_vals = ", ".join(
            [f"{name}={repr(val)}" for name, val in position_attrs.items()]
        )
        return f"Logo({arg_vals})"

    def create_image(self, sheet_type: str) -> Image:
        """Images must be created fresh in every addition.
        Using the same Image instance can cause a file write issue
        and the custom anchoring hides another issue that's more
        tricky, which is that ALL images in all tabs will anchor to
        the last anchor setting used -- unless a new image object
        is provided."""
        img = Image(self.img)
        img_size = XDRPositiveSize2D(
            cx=pixels_to_EMU(img.width), cy=pixels_to_EMU(img.height)
        )
        types = {
            "pricing": self.price_pos,
            "ratings": self.ratings_pos,
            "parts": self.parts_pos,
            "nomen-long": self.nomen_long_pos,
            "nomen-med": self.nomen_med_pos,
            "nomen-short": self.nomen_short_pos,
        }
        img.anchor = OneCellAnchor(
            _from=types[sheet_type].anchor,
            ext=img_size,
        )
        return img


class ADPLogo(Logo):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.img = str(STATIC_DIR / "adp-program-logo.png")
        self.name = "ADP Logo"


class SCALogo(Logo):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.img = str(STATIC_DIR / "sca-logo.png")
        self.name = "SCA Logo"


class CustomerLogo(Logo):
    def __init__(self, s3_key: str, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.img = S3.get_file(s3_key).file_content
        self.name = "Customer Logo"


class Logos:
    def __init__(self, *args: Logo) -> None:
        self.logos = args

    def __iter__(self):
        return iter(self.logos)

    def __repr__(self) -> str:
        return f"Logos({', '.join([repr(logo) for logo in self.logos])})"


class Cursor:
    """Track state of the cursor and provide movements
    Unpackable by `**` but replaces "col" with "column"
    as a key name"""

    def __init__(self) -> None:
        self.col = 1
        self.row = 1

    def keys(self):
        return ["row", "column"]

    def __getitem__(self, key):
        if key == "row":
            return self.row
        elif key == "column":
            return self.col

    def move_by(self, rows: int = 0, cols: int = 0) -> "Cursor":
        self.col += cols
        self.row += rows
        return self

    def move_to(self, row: int = None, col: int = None) -> "Cursor":
        self.col = col if col else self.col
        self.row = row if row else self.row
        return self

    def slam_left(self) -> "Cursor":
        self.col = 1
        return self


class PriceBook:

    def __init__(
        self,
        template: str,
        program: CustomerProgram,
    ) -> None:
        self.template_wb = opxl.load_workbook(template)
        self._9_col_template = self.template_wb["9-col"]
        self.parts_template = self.template_wb["parts-template"]
        self.product_block: CellRange = self._9_col_template["B12:J15"]
        self.parts_block: tuple[tuple[Cell]] = self.parts_template["A1:D4"]
        self.nomenclatures = opxl.load_workbook(NOMENCLATURES)
        self.program = program
        self.active_wb = self.template_wb
        self.active = self._9_col_template
        self.cursor = Cursor()
        adp_logo = ADPLogo(
            price_pos=AnchorPosition("E2", offset_x=0, offset_y=0),
            parts_pos=AnchorPosition("C2", offset_x=50, offset_y=0),
            ratings_pos=AnchorPosition("C2", offset_x=60, offset_y=0),
            nomen_long_pos=AnchorPosition("E2", offset_x=75, offset_y=0),
            nomen_med_pos=AnchorPosition("D2", offset_x=75, offset_y=0),
            nomen_short_pos=AnchorPosition("C2", offset_x=75, offset_y=0),
        )
        sca_logo = SCALogo(
            price_pos=AnchorPosition("I1", offset_x=100, offset_y=0),
            parts_pos=AnchorPosition("D1", offset_x=0, offset_y=0),
            ratings_pos=AnchorPosition("F1", offset_x=0, offset_y=0),
            nomen_long_pos=AnchorPosition("H1", offset_x=100, offset_y=0),
            nomen_med_pos=AnchorPosition("G1", offset_x=100, offset_y=0),
            nomen_short_pos=AnchorPosition("F1", offset_x=100, offset_y=0),
        )
        customer_logo = CustomerLogo(
            s3_key=program.logo_path,
            price_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
            parts_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
            ratings_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
            nomen_long_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
            nomen_med_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
            nomen_short_pos=AnchorPosition("A2", offset_x=10, offset_y=0),
        )
        self.logos = Logos(adp_logo, sca_logo, customer_logo)

    def copy_cell_style(self, src_cell: Cell, new_cell: Cell):
        # Copy font, fill, border, and alignment
        new_cell.font = copy(src_cell.font)
        new_cell.fill = copy(src_cell.fill)
        new_cell.border = copy(src_cell.border)
        new_cell.alignment = copy(src_cell.alignment)
        new_cell.number_format = src_cell.number_format
        new_cell.protection = copy(src_cell.protection)

    def copy_cell(self, src_cell: Cell, dest_row: int, dest_col: int):
        """Copy content and style from one cell to another."""
        new_cell = self.active.cell(dest_row, dest_col, src_cell.value)
        self.copy_cell_style(src_cell, new_cell)
        # NOTE below causes an error with styling indexing with
        # 'Nomenclature' tab operations.
        # so it was swapped out for the `copy_cell_style` method
        # new_cell._style = copy(src_cell._style)

    def merged_col_bounds(self, cell: Cell) -> tuple[int, int, int] | None:
        for merge_range in self.active.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
            if min_col <= cell.column <= max_col:
                if min_row <= cell.row <= max_row:
                    return min_col, max_col, min_row

    @staticmethod
    def copy_dimensions(src: Worksheet, dest: Worksheet) -> None:
        for idx, col_dim in src.column_dimensions.items():
            dest.column_dimensions[idx] = copy(col_dim)
        for idx, row_dim in src.row_dimensions.items():
            dest.row_dimensions[idx] = copy(row_dim)
        return

    def active_cell(self, value=None) -> Cell:
        return self.active.cell(**self.cursor, value=value)

    def new_sheet(
        self,
        template: Worksheet,
        title: str,
        sheet_type: str,
        include_logos: bool = True,
    ) -> "PriceBook":
        new_sheet = self.active_wb.copy_worksheet(template)
        new_sheet.title = title.strip()
        self.active = new_sheet
        if include_logos:
            for logo in self.logos:
                try:
                    logo_img = logo.create_image(sheet_type=sheet_type)
                    self.active.add_image(logo_img, logo_img.anchor)
                except Exception as e:
                    logger.warning(f"Logo insertion failed for {logo}: {e}")
        self.active.sheet_view.showGridLines = False
        self.cursor.move_to(1, 1)
        return self

    def new_price_sheet(self, title: str) -> "PriceBook":
        sheet_type = "pricing"
        new_sheet_template = self._9_col_template
        (self.min_col, self.min_row, self.max_col, self.max_row) = range_boundaries(
            "B12:L15"
        )
        new_sheet = self.new_sheet(
            template=new_sheet_template, title=title, sheet_type=sheet_type
        )
        self.cursor.move_to(1, 1)
        return new_sheet

    def new_ratings_sheet(self, title: str = "M1 Ratings") -> "PriceBook":
        sheet_type = "ratings"
        template = self.template_wb["ratings"]
        return self.new_sheet(
            template=template, title=title, sheet_type=sheet_type, include_logos=False
        )

    def new_parts_sheet(self) -> "PriceBook":
        sheet_type = "parts"
        template = self.parts_template
        title = "Parts"
        return self.new_sheet(template=template, title=title, sheet_type=sheet_type)

    def new_nomenclature_sheet(self, len_option: LenOption) -> "PriceBook":
        sheet_type = f"nomen-{len_option}"
        template = self.template_wb[sheet_type]
        title = "Nomenclature"
        return self.new_sheet(
            template=template, title=title, sheet_type=sheet_type, include_logos=False
        )

    def movex(self, n_cols: int) -> "PriceBook":
        self.cursor.move_by(cols=n_cols)
        return self

    def movey(self, n_rows: int) -> "PriceBook":
        self.cursor.move_by(rows=n_rows)
        return self

    def insert_rows(self, n_rows: int = 1) -> "PriceBook":
        starting_row = self.cursor.row
        last_current_row = self.active[self.cursor.row]
        self.active.insert_rows(self.cursor.row, amount=n_rows)
        for _ in range(1, n_rows + 1):
            for col, cell in enumerate(self.active[self.cursor.row]):
                cell._style = copy(last_current_row[col]._style)
            self.cursor.move_by(rows=1)
        self.cursor.move_to(row=starting_row)
        return self

    def append_blank_product_block(self, col_offset: int = 0) -> "PriceBook":
        self.cursor.slam_left().move_by(rows=3, cols=col_offset)
        row_offset = self.cursor.row - self.min_row
        ## copy new product block from template
        for row in self.product_block:
            for cell in row:
                new_row = cell.row + row_offset
                new_col = cell.column
                self.copy_cell(cell, new_row, new_col)
        self.active.merge_cells(
            start_row=self.cursor.row,
            start_column=self.cursor.col,
            end_row=self.cursor.row,
            end_column=len(self.product_block[0]) + self.cursor.col - 1,
        )
        self.cursor.slam_left()
        return self

    def append_parts_block(self, col_offset: int = 0) -> "PriceBook":
        self.cursor.slam_left().move_by(cols=col_offset)
        min_col, min_row, max_col, max_row = range_boundaries("A1:D4")
        row_offset = self.cursor.row - min_row
        for row in self.parts_block:
            for cell in row:
                new_row = cell.row + row_offset
                new_col = cell.column + col_offset
                self.copy_cell(cell, new_row, new_col)
        self.active.merge_cells(
            start_row=self.cursor.row,
            start_column=self.cursor.col,
            end_row=self.cursor.row,
            end_column=len(self.parts_block[0]) + self.cursor.col - 1,
        )
        self.cursor.slam_left()
        return self

    def find_min_coords_with_data(self, sheet: Worksheet = None) -> MinCoords:
        if not sheet:
            sheet = self.active
        min_row = sheet.max_row
        min_col = sheet.max_column
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    min_row = min(min_row, cell.row)
                    min_col = min(min_col, cell.column)
        if min_row == sheet.max_row:
            min_row = None
        if min_col == sheet.max_column:
            min_col = None
        return min_row, min_col

    def insert_nomenclature_block(
        self, series: str, offset: tuple[int, int] = (0, 0)
    ) -> "PriceBook":
        nomenclature_sheet: Worksheet = self.nomenclatures[series]
        (model_type,) = tuple([e for e in MODELS if e.__name__ == series])
        model_example = self.program.sample_from_program(series=series)
        pl_remapping: dict[ModelSeries, ModelSeries] = {
            HE: CE,
            V: CE,
            MH: CE,
            B: CF,
            S: CF,
            F: CF,
        }
        try:
            if result := re.match(model_type.regex, model_example, re.VERBOSE):
                model_nomenclature = result.groupdict()
            else:
                result = re.match(
                    pl_remapping[model_type].regex, model_example, re.VERBOSE
                )
                if not result:
                    raise Exception("Could not match model")
                model_nomenclature = result.groupdict()

            if model_nomenclature.get("revision"):
                model_nomenclature.pop("rds", None)
            elif model_nomenclature.get("rds"):
                model_nomenclature.pop("revision", None)
            ignore_custom_model_insertion = False
        except Exception as e:
            logger.warning(
                f"Error with sample {model_example}. Nomenclature will show its default"
            )
            ignore_custom_model_insertion = True
        else:
            if "scode" in model_nomenclature and "mat" in model_nomenclature:
                if series == "CP":
                    model_nomenclature["scode"] = (
                        model_nomenclature["scode"] + model_nomenclature["mat"]
                    )
                else:
                    model_nomenclature["scode"] = (
                        model_nomenclature["mat"] + model_nomenclature["scode"]
                    )
                model_nomenclature.pop("mat")
        for i, row in enumerate(nomenclature_sheet):
            row_num = i + 1
            for cell in row:
                new_row = cell.row + self.cursor.row - 1
                new_col = cell.column + self.cursor.col - 1
                self.copy_cell(cell, new_row, new_col)
            if row_num == 2 and not ignore_custom_model_insertion:
                self.cursor.move_by(1, 1)
                for val in model_nomenclature.values():
                    self.active_cell(value=val)
                    self.cursor.move_by(cols=1)
                self.cursor.move_by(-1, -1).slam_left().move_by(*offset)
        _, min_col = self.find_min_coords_with_data(nomenclature_sheet)
        self.active.merge_cells(
            start_row=self.cursor.row,
            start_column=min_col,
            end_row=self.cursor.row,
            end_column=nomenclature_sheet.max_column,
        )
        self.cursor.move_by(rows=nomenclature_sheet.max_row).slam_left().move_by(
            *offset
        )
        return self

    def attach_nomenclature_tab(self) -> "PriceBook":
        long_nomens = {"HE", "HH", "B", "F", "CP", "LS"}
        medium_nomens = {"V", "CE", "CF"}
        if self.program.series_contained.intersection(long_nomens):
            self.new_nomenclature_sheet("long").movey(1)
        elif self.program.series_contained.intersection(medium_nomens):
            self.new_nomenclature_sheet("med").movey(1)
        else:
            self.new_nomenclature_sheet("short").movey(1)

        for series in self.program.series_contained:
            self.insert_nomenclature_block(series).movey(2)
        for col in self.active.iter_cols():
            col_letter = col[0].column_letter
            active_col = self.active.column_dimensions[col_letter]
            active_col.width = NOMENCLATURE_COL_WIDTH
        return self

    def save_and_close(self) -> BytesIO:
        file_obj = BytesIO()
        try:
            self.active_wb.remove(self._9_col_template)
            self.active_wb.remove(self.template_wb["ratings"])
            for option in ("long", "med", "short"):
                self.active_wb.remove(self.template_wb[f"nomen-{option}"])
            self.active_wb.remove(self.parts_template)
            # set the opening sheet to the first sheet,
            # which should be the model list
            self.active_wb.active = 0
            self.active_wb.save(file_obj)
        except IndexError as e:
            import traceback as tb

            logger.critical("file empty on save for " f"{self.program.customer_name}")
            logger.critical(tb.format_exc())
            raise FileGenExecption
        finally:
            self.template_wb.close()
        file_obj.seek(0)
        return file_obj

    def add_footer(self, offset: tuple[int, int] = (0, 0)) -> "PriceBook":
        self.cursor.move_by(*offset)
        for name, value in self.program.terms.items():
            cell = self.active_cell(name + ": ")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")
            self.movex(1)
            cell = self.active_cell(value["value"])
            for style_name, style_value in value["style"].items():
                cell.__setattr__(style_name, style_value)
                cell.alignment = Alignment(horizontal="left")
            self.cursor.slam_left().move_by(rows=1).move_by(*offset)
        return self

    def insert_data(
        self, df: pd.DataFrame, headers: bool = True, offset: tuple = (0, 0)
    ) -> "PriceBook":
        self.cursor.move_by(*offset)
        if headers:
            for col in df:
                if col == Fields.STAGE.formatted():
                    pass
                else:
                    self.active_cell(value=str(col))
                self.cursor.move_by(cols=1)
            self.cursor.slam_left().move_by(rows=1).move_by(*offset)

        for label, data in df.iterrows():
            for datum in data:
                if datum == Stage.PROPOSED.name:
                    cell = self.active_cell(value=datum.lower())
                    cell.font = Font(italic=True, size=8.0, color="808080")
                elif datum == Stage.ACTIVE.name:
                    self.active_cell(value=None)
                else:
                    self.active_cell(value=datum)
                self.cursor.move_by(cols=1)
            self.cursor.slam_left().move_by(rows=1).move_by(*offset)
        if headers:
            if Fields.PALLET_QTY.formatted() in df.columns:
                disclaimer_txt = "* Must order in pallet quantities"
                self.cursor.move_by(cols=1)
                disclaimer = self.active_cell(disclaimer_txt)
                disclaimer.font = Font(bold=True, italic=True)
            elif Fields.MIN_QTY.formatted() in df.columns:
                disclaimer_txt = "* Must order at least the minimum quantity per-SKU"
                self.cursor.move_by(cols=1)
                disclaimer = self.active_cell(disclaimer_txt)
                disclaimer.font = Font(bold=True, italic=True)
            self.cursor.slam_left()
        return self

    @staticmethod
    def split_rated_from_unrated(category_data: DataByCategory) -> DataByCategory:
        result = dict()
        for category, df in category_data.items():
            all_rated = df["rated"].all()
            all_unrated = (~df["rated"]).all()
            new_category_rated = category + " - Rated"
            new_category_unrated = category + " - Replacement"
            if all_rated:
                result[new_category_rated] = df
            elif all_unrated:
                result[new_category_unrated] = df
            else:
                result[new_category_rated] = df[df["rated"]]
                result[new_category_unrated] = df[~df["rated"]]
        return result

    def build_program(self, session: Session) -> "PriceBook":
        """
        One file per "program"
        One product block per "category"
        """
        offset = (0, 1)
        self.new_price_sheet(title="Model List")
        self.cursor.move_by(11).slam_left()
        for i, program in enumerate(self.program):
            data_by_category: dict[str, pd.DataFrame] = {
                category: program.category_data(
                    category, self.program.customer_id, session
                )
                for category in program.product_categories
            }
            data_by_category = self.split_rated_from_unrated(data_by_category)
            for j, cat_data in enumerate(data_by_category.items()):
                cat, df = cat_data
                df_copy = df.copy(deep=True)
                df_copy.drop(
                    columns=[Fields.RATED.value, Fields.SERIES.value], inplace=True
                )
                df_copy.rename(
                    columns={col: Fields(col).formatted() for col in df_copy.columns},
                    inplace=True,
                )
                cat = re.sub(r"(- Embossed )|( Painted)", "", cat)
                rows = df_copy.shape[0]
                if i > 0 or j > 0:
                    self.append_blank_product_block(col_offset=1)
                match offset:
                    case 0, 0:
                        self.active_cell(value=cat)  # put in category
                    case _:
                        self.cursor.move_by(*offset)
                        self.active_cell(value=cat)  # put in category
                        self.cursor.slam_left()
                (
                    self.movey(2)  # the first data row
                    .adjust_number_of_formatted_rows(
                        self.max_row - self.min_row - 1, rows
                    )
                    .movey(-1)  # headers row
                    .insert_data(df_copy, offset=offset)
                )
        return self.movex(-self.cursor.col + 1).movey(2).attach_parts(offset=offset)

    def adjust_number_of_formatted_rows(
        self, startin_row_num: int, target_row_num: int
    ) -> "PriceBook":
        diff = target_row_num - startin_row_num
        if diff > 0:
            self.insert_rows(n_rows=diff)
        elif diff < 0:
            self.active.delete_rows(self.cursor.row, amount=abs(diff))
        return self

    def attach_parts(self, offset: tuple[int, int] = (0, 0)) -> "PriceBook":
        """unlike the model numbers tab, headers have been provided"""
        data = self.program.parts
        if data.empty:
            logger.info("skipped parts due to empty table")
            return self
        num_rows = data.shape[0]
        self.append_parts_block(col_offset=offset[1]).movey(
            2
        ).adjust_number_of_formatted_rows(2, num_rows).insert_data(
            data, headers=False, offset=offset
        )
        self.cursor.slam_left()
        return self.movey(1)

    def attach_ratings(self) -> "PriceBook":
        data = self.program.ratings
        if data.empty:
            return self

        def set_series_name(oemseries: str) -> str:
            shorter_prefixes: list = loads(os.getenv("SHORT_OEM_PREFIXES", "[]"))
            longer_prefixes: list = loads(os.getenv("LONG_OEM_PREFIXES", "[]"))
            if any([val in oemseries for val in longer_prefixes]):
                start, end = (0, 5)
            elif any(oemseries.startswith(prefix) for prefix in shorter_prefixes):
                start, end = (0, 3)
            else:
                start, end = (0, 4)
            return oemseries[start:end]

        data["OEM Series"] = (
            data["OEM Series"].fillna(data["OutdoorModel"]).apply(set_series_name)
        )
        data["OEM Name"] = data["OEM Name"].fillna(data["OEMName"])
        ratings: dict[str, pd.DataFrame] = {
            oem_series: data[data["OEM Series"] == oem_series]
            for oem_series in data["OEM Series"].unique()
        }
        substituted_series = []
        new_series = dict()
        for series, data in ratings.items():
            unique_oems = data["OEM Name"].unique().tolist()
            if len(unique_oems) > 1:
                for oem in unique_oems:
                    tab_name = f"{series} {oem}"
                    new_series[tab_name] = data[data["OEM Name"] == oem]
                substituted_series.append(series)
        if substituted_series:
            ratings |= new_series
            for series in substituted_series:
                del ratings[series]

        for tab, table in ratings.items():
            if len(tab) > 31:
                tab = tab[:31].strip()
            if (
                table["FurnaceModel"].isna().all()
                and table["Furnace Model Number"].isna().all()
            ):
                include_furnace_col = False
            else:
                include_furnace_col = True

            if table["HSPF2"].isna().all() or (table["HSPF2"] == 0).all():
                include_HSPF_col = False
            else:
                include_HSPF_col = True

            rows = table.shape[0]
            self.new_ratings_sheet(title=tab)
            if not include_HSPF_col:
                self.active.delete_cols(9, 1)
            if not include_furnace_col:
                self.active.delete_cols(5, 1)
                self.active.column_dimensions["E"].width = 10.5
            self.cursor.move_to(row=3)  # second row for data
            self.adjust_number_of_formatted_rows(3, rows)
            self.cursor.move_by(rows=-1)  # back to first data row

            # column names are already in the template
            # rearrange data to align with the template and copy it in
            # template_cols = ('AHRINumber','OEMName','OutdoorModel',
            #                 'IndoorModel','FurnaceModel',
            #                 'SEER2','EER95F2',
            #                 'Capacity2','HSPF2')

            table["Model Number"].fillna(table["OutdoorModel"], inplace=True)
            table["Coil Model Number"].fillna(table["IndoorModel"], inplace=True)
            table["ADP Series"].ffill(inplace=True)
            table = table.sort_values(
                by=["ADP Series", "Model Number", "Coil Model Number", "OEMName"],
                ascending=True,
            )

            for label, row in table.iterrows():
                if not row["AHRI Ref Number"]:
                    row_view = row[
                        [
                            "AHRINumber",
                            "OEMName",
                            "OutdoorModel",
                            "IndoorModel",
                            "FurnaceModel",
                            "seer2_as_submitted",
                            "eer95f2_as_submitted",
                            "capacity2_as_submitted",
                            "hspf2_as_submitted",
                        ]
                    ]
                    if not include_furnace_col:
                        row_view = row_view.drop(index=["FurnaceModel"])
                    if not include_HSPF_col:
                        row_view = row_view.drop(index=["hspf2_as_submitted"])
                else:
                    row_view = row[
                        [
                            "AHRI Ref Number",
                            "OEM Name",
                            "Model Number",
                            "Coil Model Number",
                            "Furnace Model Number",
                            "SEER2",
                            "EER2",
                            "Capacity2",
                            "HSPF2",
                        ]
                    ]
                    if not include_furnace_col:
                        row_view = row_view.drop(index=["Furnace Model Number"])
                    if not include_HSPF_col:
                        row_view = row_view.drop(index=["HSPF2"])
                for datum in row_view:
                    cell = self.active_cell(value=datum)
                    if datum == "0":
                        cell.value = "pending"
                        cell.font = Font(italic=True)
                    elif not datum:
                        cell.value = ""
                    self.cursor.move_by(cols=1)
                self.cursor.slam_left()
                self.cursor.move_by(rows=1)
        return self
